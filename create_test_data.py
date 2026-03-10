"""Generate dummy CM test data for M1 dashboard testing."""
import openpyxl
import os
from datetime import datetime, timedelta

sync_path = r'C:\Users\ajamitho\OneDrive - amazon.com\Model 2 Test - JT-TEST'

test_cms = [
    {"alias": "CM1_StarPerformer", "stage": "Stage 3 — Full Partnership", "checkins": 12, "scenarios": "6 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=28)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.5, "scores": [3,4,4,3,3,3,4,4]},
         {"date": (datetime.now()-timedelta(days=21)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.8, "scores": [4,4,4,3,4,3,4,4]},
         {"date": (datetime.now()-timedelta(days=14)).strftime("%Y-%m-%d"), "stage": "Stage 3 — Full Partnership", "avg": 4.1, "scores": [4,4,5,4,4,3,4,5]},
         {"date": (datetime.now()-timedelta(days=2)).strftime("%Y-%m-%d"), "stage": "Stage 3 — Full Partnership", "avg": 4.5, "scores": [5,4,5,4,5,4,4,5]},
     ],
     "agenda": ["Discuss Stage 3 ownership goals"],
     "milestones": [("Shadow CCM on 3 campaigns","Complete","Stage 1"),("Run campaign independently","Complete","Stage 2"),("Mentor new CM","In Progress","Stage 3")]},
    {"alias": "CM2_SteadyProgress", "stage": "Stage 2 — Becoming Independent", "checkins": 8, "scenarios": "5 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=21)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.5, "scores": [2,3,3,2,2,3,2,3]},
         {"date": (datetime.now()-timedelta(days=14)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.0, "scores": [3,3,3,3,3,3,3,3]},
         {"date": (datetime.now()-timedelta(days=5)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.4, "scores": [3,4,4,3,3,3,3,4]},
     ],
     "agenda": [],
     "milestones": [("Shadow CCM on 3 campaigns","Complete","Stage 1"),("Run campaign independently","In Progress","Stage 2")]},
    {"alias": "CM3_ConfidenceDrop", "stage": "Stage 2 — Becoming Independent", "checkins": 6, "scenarios": "4 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=14)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.5, "scores": [4,3,4,3,3,4,3,4]},
         {"date": (datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.8, "scores": [4,4,4,3,4,4,3,4]},
         {"date": (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 2.6, "scores": [2,3,3,2,3,2,3,2]},
     ],
     "agenda": ["Review trafficking workflow confusion","Clarify escalation paths"],
     "milestones": [("Shadow CCM on 3 campaigns","Complete","Stage 1"),("Run campaign independently","In Progress","Stage 2")]},
    {"alias": "CM4_Inactive", "stage": "Stage 1 — Building Trust", "checkins": 2, "scenarios": "1 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=18)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.0, "scores": [2,2,2,2,2,2,2,2]},
         {"date": (datetime.now()-timedelta(days=12)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.1, "scores": [2,2,2,2,2,2,3,2]},
     ],
     "agenda": ["Need help understanding campaign setup"],
     "milestones": [("Shadow CCM on 3 campaigns","In Progress","Stage 1")]},
    {"alias": "CM5_NewStarter", "stage": "Stage 1 — Building Trust", "checkins": 1, "scenarios": "0 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=3)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 1.5, "scores": [1,2,2,1,1,1,2,2]},
     ],
     "agenda": [],
     "milestones": [("Shadow CCM on 3 campaigns","Not Started","Stage 1")]},
    {"alias": "CM6_StrongMidStage", "stage": "Stage 2 — Becoming Independent", "checkins": 9, "scenarios": "6 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=21)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.8, "scores": [3,3,3,2,3,2,3,3]},
         {"date": (datetime.now()-timedelta(days=14)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.5, "scores": [4,3,4,3,3,4,3,4]},
         {"date": (datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.9, "scores": [4,4,4,4,4,3,4,4]},
         {"date": (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 4.0, "scores": [4,4,4,4,4,4,4,4]},
     ],
     "agenda": ["Review Stage 3 readiness"],
     "milestones": [("Shadow CCM on 3 campaigns","Complete","Stage 1"),("Run campaign independently","Complete","Stage 2"),("Mentor new CM","Not Started","Stage 3")]},
    {"alias": "CM7_Struggling", "stage": "Stage 1 — Building Trust", "checkins": 4, "scenarios": "2 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=21)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.0, "scores": [2,2,2,2,2,2,2,2]},
         {"date": (datetime.now()-timedelta(days=14)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 1.8, "scores": [2,1,2,2,2,1,2,2]},
         {"date": (datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 1.6, "scores": [1,1,2,2,1,1,2,2]},
         {"date": (datetime.now()-timedelta(days=2)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 1.5, "scores": [1,1,2,1,1,1,2,2]},
     ],
     "agenda": ["Basic campaign setup help","Understanding escalation process","Struggling with prioritization"],
     "milestones": [("Shadow CCM on 3 campaigns","In Progress","Stage 1")]},
    {"alias": "CM8_StaleAgenda", "stage": "Stage 2 — Becoming Independent", "checkins": 5, "scenarios": "3 / 6",
     "history": [
         {"date": (datetime.now()-timedelta(days=21)).strftime("%Y-%m-%d"), "stage": "Stage 1 — Building Trust", "avg": 2.5, "scores": [2,3,3,2,2,3,2,3]},
         {"date": (datetime.now()-timedelta(days=10)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.0, "scores": [3,3,3,3,3,3,3,3]},
         {"date": (datetime.now()-timedelta(days=8)).strftime("%Y-%m-%d"), "stage": "Stage 2 — Becoming Independent", "avg": 3.1, "scores": [3,3,4,3,3,3,3,3]},
     ],
     "agenda": ["Pending: Review QA checklist (from 3 weeks ago)","Pending: Discuss optimization strategy"],
     "milestones": [("Shadow CCM on 3 campaigns","Complete","Stage 1"),("Run campaign independently","In Progress","Stage 2")]},
]

dim_headers = ["Trafficking","Optimizations","Communication","Prioritization","Navigation","Ambiguity","Partnership","Independence"]
created = 0

for cm in test_cms:
    cm_dir = os.path.join(sync_path, cm["alias"])
    os.makedirs(cm_dir, exist_ok=True)
    fp = os.path.join(cm_dir, "cm_progress.xlsx")
    wb = openpyxl.Workbook()
    
    # Sheet 1: Current Snapshot
    ws1 = wb.active
    ws1.title = "Current Snapshot"
    ws1.append(["Field", "Value"])
    last_h = cm["history"][-1]
    ws1.append(["Current Stage", cm["stage"]])
    ws1.append(["Overall Confidence", f'{last_h["avg"]:.1f} / 5.0'])
    ws1.append(["Last Updated", last_h["date"]])
    ws1.append(["Total Check-ins", cm["checkins"]])
    ws1.append(["Scenarios Practiced", cm["scenarios"]])
    
    # Sheet 2: Confidence History
    ws2 = wb.create_sheet("Confidence History")
    ws2.append(["Date", "Stage", "Average"] + dim_headers)
    for h in cm["history"]:
        ws2.append([h["date"], h["stage"], h["avg"]] + h["scores"])
    
    # Sheet 3: M1 Agenda Topics
    ws3 = wb.create_sheet("M1 Agenda Topics")
    ws3.append(["Topic"])
    for t in cm["agenda"]:
        ws3.append([t])
    
    # Sheet 4: Milestones
    ws4 = wb.create_sheet("Milestones")
    ws4.append(["Milestone", "Status", "Stage"])
    for mn, ms, mst in cm["milestones"]:
        ws4.append([mn, ms, mst])
    
    wb.save(fp)
    wb.close()
    created += 1
    print(f"  OK: {cm['alias']}")

print(f"\nDone: {created}/8 CM folders created in {sync_path}")
